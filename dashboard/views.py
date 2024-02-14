from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from main import models
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from itertools import chain
from datetime import datetime
from . funcs import search_with_fields, pagenator_page
from django.core.exceptions import FieldError


def dashboard(request):
    categorys = models.Category.objects.all()
    products = models.Product.objects.filter(is_active=True)
    users = User.objects.filter(is_satff=False)
    context = {
        'categorys':categorys,
        'products':products,
        'users':users,
    }
    return render(request, 'index.html', context)


def category_list(request):
    categorys = models.Category.objects.all()
    return render(request, 'category/list.html', {'categorys':categorys})


def category_detail(request, id):
    category = models.Category.objects.get(id=id)
    products = models.Product.objects.filter(category=category, is_active=True)
    context = {
        'category':category,
        'products':products
    }
    return render(request, 'category/list.html', context)


def category_update(request, id):
    category = models.Category.objects.get(id=id)
    category.name = request.POST['name']
    category.save()
    return redirect('category_detail', category.id)


def category_delete(request, id):
    category = models.Category.objects.get(id=id)
    category.delete()
    return redirect('category_list')



# CRUD


# dashboard
@login_required(login_url='sign_in')
def dashboard(request):
    
    users = User.objects.all().count()
    products = models.Product.objects.filter().count()
    category = models.Category.objects.all().count()

    context = {
        'users':users,
        'products':products,
        'category':category
    }

    return render(request, 'dashb/index.html', context)


#                                                       category CRUD
def create_category(request):
    if request.method == 'POST':
        models.Category.objects.create(
            name=request.POST['name']
        )
        return redirect('categorys')
    return render(request, 'dashb/category/create.html')


def categorys(request):
    categorys = models.Category.objects.all()
    return render(request, 'dashb/category/list.html', {'categorys':categorys})


def category_update(request, id):
    category = models.Category.objects.get(id=id)
    if request.method == 'POST':
        category.name = request.POST['name']
        category.save()

    return render(request, 'dashb/category/update.html', {'category':category})


def category_delete(request, id):
    models.Category.objects.get(id=id).delete()
    return redirect('category')


#product

#def productss(request):
#    products = models.Product.objects.all()
#    return render(request, 'dashb/items/lists.html', {'products':products})
def productss(request):
    try:
        result = search_with_fields(request)
        products = models.EnterProduct.objects.filter(**result)
    except FieldError as err:
        field_name = err.__doc__.split()[3][1:-1]
        if field_name in result:
            del result[field_name]
            products = models.Product.objects.filter(**result)
        else:
            # Field is not found in result, handle this situation accordingly
            # or raise an error if needed
            products = models.Product.objects.none()

    context = {'products': pagenator_page(products, 1, request)}
    return render(request, 'dashb/items/lists.html', context)


def product_create(request):
    categorys = models.Category.objects.all()
    context = {
        'categorys':categorys
    }
    if request.method == "POST":
        name = request.POST['name']
        description = request.POST['description']
        quantity = request.POST['quantity']
        price = request.POST['price']
        currency = request.POST['currency']
        baner_image = request.FILES['baner_image']
        category_id = request.POST['category_id']
        images = request.FILES.getlist('images')
        product = models.Product.objects.create(
            name=name,
            description = description,
            quantity=quantity,
            price=price,
            currency=currency,
            baner_image=baner_image,
            category_id=category_id
        )
        for image in images:
            models.ProductImage.objects.create(
                image=image,
                product=product
            )

    return render(request, 'dashb/items/create.html', context)


#def products(request):
#    name = request.GET.get('name')
#    if name:
#        products = models.EnterProduct.objects.filter(
#            product__name=name,
#        )
#    else:
#        products = models.Product.objects.all()
#    context  = {
#        'products':products,
#
#    }
#    return render(request, 'dashb/items/list.html', context)


def products(request):
    try:
        result = search_with_fields(request)
        products = models.EnterProduct.objects.filter(**result)
    except FieldError as err:
        field_name = err.__doc__.split()[3][1:-1]
        if field_name in result:
            del result[field_name]
            products = models.Product.objects.filter(**result)
        else:
            # Field is not found in result, handle this situation accordingly
            # or raise an error if needed
            products = models.Product.objects.none()

    context = {'products': pagenator_page(products, 1, request)}
    return render(request, 'dashb/items/list.html', context)



def product_update(request, id):
    product = models.Product.objects.get(id=id)
    if request.method == 'POST':
        category = models.Category.objects.get(id=request.POST['category_id'])
        product.name = request.POST['name']
        product.description=request.POST['description']
        product.quantity = request.POST['quantity']
        product.price = request.POST['price']
        product.currency = request.POST['currency']
        product.discount_price = request.POST['discount_price']
        product.description = request.POST['description']
        product.category=category
        baner_image = request.FILES.get('baner_image')
        if baner_image:
            product.baner_image=baner_image
        product.save()

    context  = {
        'product':product,
        'categorys':models.Category.objects.all(),
    }

    return render(request, 'dashb/items/update.html', context)


def product_delete(request, id):
    models.Product.objects.get(id=id).delete()
    return redirect('items')


#auth

def register_user(request):
    if request.method == 'POST':
        User.objects.create_user(
            username = request.POST['username'],
            password = request.POST['password']
        )

    return render(request, 'dashb/auth/register.html')


def sign_in(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashb')
    return render(request, 'dashb/auth/login.html')


def sign_out(request):
    logout(request)
    return redirect('index')


def product_detail(request, id):
    product = models.Product.objects.get(id=id)
    enters = models.EnterProduct.objects.filter(product=product)
    outs = models.CartProduct.objects.filter(product=product, card__is_active=False)
    query_set = sorted(
        chain(
            enters,
            outs
        ),
        key = lambda x : x.created_at
    )
    # for i in query_set:
    #     try:
    #         i.card
    #         print(f"chiqish {i}")
    #     except:
    #         print(f"kirish {i}")

    return render(request, 'dashb/items/detail.html', {'query_set':query_set})

def create_enter(request):
    if request.method == 'POST':
        product_id = request.POST['product_id']
        quantity = int(request.POST['quantity'])
        models.EnterProduct.objects.create(
            product_id=product_id,
            quantity=quantity
        )
        return redirect('list_enter')
    return render(request, 'dashb/enter/create.html', {'products':models.Product.objects.all()})


def update_enter(request, id):
    if request.method == 'POST':
        quantity = int(request.POST['quantity'])
        enter = models.EnterProduct.objects.get(id=id)
        enter.quantity = quantity
        enter.save()
    return redirect('dashb:list_enter')


def delete_enter(request, id):
    models.EnterProduct.objects.get(id=id).delete()
    return redirect('dashb:list_enter')

#def list_enter(request):
#    result = search_with_fields(request)
#    try: 
#        enters = models.EnterProduct.objects.filter(**result)
#    except FieldError as err:
#        del result[err.__doc__.split()[3][1:-1]]
#        enters = models.EnterProduct.objects.filter(**result)
#
#    context = {'enters': pagenator_page(enters, 1, request)}
#    return render(request, 'dashb/enter/list.html', context)



def list_enter(request):
    result = search_with_fields(request)
    try: 
        enters = models.EnterProduct.objects.filter(**result)
    except FieldError as err:
        field_name = err.__doc__.split()[3][1:-1]
        if field_name in result:
            del result[field_name]
            enters = models.EnterProduct.objects.filter(**result)
        else:
            # Field is not found in result, handle this situation accordingly
            # or raise an error if needed
            enters = models.EnterProduct.objects.none()

    context = {'enters': pagenator_page(enters, 1, request)}
    return render(request, 'dashb/enter/list.html', context)



#def list_enter(request):
#    result = search_with_fields(request)
#    try: 
#        enters = models.EnterProduct.objects.filter(**result)
#
#    except FieldError as err:
#        del result[err.__doc__.split()[3][1:-1]]
#        enters = models.EnterProduct.objects.filter(**result)
#
#    context = {'enters': pagenator_page(enters, 1, request)}
#    return render(request, 'dashb/enter/list.html', context)




#def list_enter(request):
#    name = request.GET.get('name')
#    quantity = request.GET.get('quantity')
#    created_at = request.GET.get('created_at')
#    if name and quantity and created_at:
#        enters = models.EnterProduct.objects.filter(
#            product__name=name,
#            quantity=quantity,
#            created_at__gt = datetime.strptime(created_at, '%Y-%m-%dT%H:%M'),
#            created_at__lte = datetime.strptime(created_at, '%Y-%m-%dT%H:%M'),
#        )
#    else:
#        enters = models.EnterProduct.objects.all()
#    context = {'enters':enters}
#    return render(request, 'dashb/enter/list.html', context)
#

# views.py

def kirim(request):
    kirims = models.EnterProduct.objects.all()
    products = models.Product.objects.all()
    context = {'kirims':kirims,
               'products':products}
    return render(request, 'dashb/items/kirim.html', context)





def generate_excel(request):
    # Ma'lumotlarni olish
    enters = models.EnterProduct.objects.all()

    # Excel faylni yaratish
    wb = Workbook()
    ws = wb.active

    # Ma'lumotlarni Excel fayliga yozish
    ws.append(['№', 'Maxsulot nomi', 'Soni', 'Sana'])
    
    for enter in enters:
        row = [
            enter.id,
            enter.product.name if enter.product else enter.product_name,
            enter.quantity,
            enter.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        ws.append(row)


    # Response obyektini yaratish
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="malumotlar.xlsx"'

    # Excel faylni HttpResponse ga yozish
    wb.save(response)

    return response







from openpyxl import load_workbook

def import_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        
        # Excel faylini yuklash
        wb = load_workbook(excel_file)
        ws = wb.active

        # Ma'lumotlarni bazaga saqlash
        for row in ws.iter_rows(min_row=2, values_only=True):  # 1-qator nomlari hisobga olinmaydi, shuning uchun 2-qatordan boshlaymiz
            product_name = row[1]
            quantity = row[2] or 0  # Agar quantity bo'sh bo'lsa, 0 deb qo'yamiz
            created_at = row[3]

            # Sana qiymati string ko'rinishida, uni datetime obyektiga aylantiramiz
            created_at = datetime.strptime(created_at, '%Y-%m-%d %H:%M')

            enter_product = models.EnterProduct.objects.create(
                product_name=product_name,
                quantity=quantity,
                created_at=created_at
            )

        # Sahifani qayta yuklash uchun redirect qiling
        return HttpResponse("Excel fayl muvaffaqiyatli saqlandi!")
    else:
        return render(request, 'your_template_name.html')
